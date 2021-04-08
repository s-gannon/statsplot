from matplotlib import pyplot as plt
import openpyxl as xl
import numpy as np
import math
import re   #regex ftw

"""def resid(x_pts, y_pts, m, b):  #returns the new y points to be plotted (as resid)
    resid_pts = []
    for x in range(len(y_pts)):
        resid_pts.append(y_pts[x] - ((m * x_pts[x]) + b))
    print(resid_pts)
    return resid_pts
"""
#   function declarations
def getR(x_pts, y_pts):         #gets the correlation value
    mean_x = mean_y = 0
    for x in range(len(x_pts)):
        mean_x += x_pts[x]
        mean_y += y_pts[x]
    mean_x = mean_x/len(x_pts)
    mean_y = mean_y/len(y_pts)
    a = x_pts
    b = y_pts
    for x in range(len(x_pts)):
        a[x] = mean_x - x_pts[x]
        b[x] = mean_y - y_pts[x]
    ab = a_squared = b_squared = []
    for x in range(len(a)):
        ab.append(a[x]*b[x])
        a_squared.append(a[x]*a[x])
        b_squared.append(b[x]*b[x])
    sum_ab = sum_a_sq = sum_b_sq = 0
    for x in range(len(ab)):
        sum_ab += ab[x]
        sum_a_sq += a_squared[x]
        sum_b_sq += b_squared[x]
    return (float(sum_ab)/math.sqrt(float(sum_a_sq) * float(sum_b_sq)))
#   variable initialization
x_points = []
y_points = []
x_range = ["",[0,0]]
y_range = ["",[0,0]]
wb = ws = ""
#   Main
filename = input("Enter in the Excel file name: ")
while re.match(r"^([A-Z]:\\((\w+\\)+)?)?([\w\-_]+.xlsx)$", filename) == None:
    print("Invalid file name. Check your path and file extension.")
    filename = input("Enter in the Excel file name: ")
try:
    wb = xl.load_workbook(filename)
    ws = wb.active
except:
    print("An exception occured loading the workbook and worksheet.")
    print("Check to make sure your path/filename is spelled correctly")
    exit()
#getting input for the start & end of the x & y data
x_start = str(input("Enter the start of the set of data for x-axis (format: A1): ")).upper()
x_end = str(input("Enter the end of the set of data for x-axis (format: A1): ")).upper()
y_start = str(input("Enter the start of the set of data for y-axis (format: A1): ")).upper()
y_end = str(input("Enter the end of the set of data for y-axis (format: A1): ")).upper()

x_range[0] = x_start[0]
x_start = x_start.replace(x_start[0], "")
x_end = x_end.replace(x_end[0], "")
x_range[1] = [int(x_start), int(x_end)]

y_range[0] = y_start[0]
y_start = y_start.replace(y_start[0], "")
y_end = y_end.replace(y_end[0], "")
y_range[1] = [int(y_start), int(y_end)]

for x in range(x_range[1][0], x_range[1][1] + 1):   #min start at 1, goes to n
    x_points.append(str(x_range[0] + str(x)))
    
for y in range(y_range[1][0], y_range[1][1] + 1):   #min start at 1, goes to n
    y_points.append(str(y_range[0] + str(y)))

for x in range(len(x_points)):
    x_points[x] = ws[x_points[x]].value

for y in range(len(y_points)):
    y_points[y] = ws[y_points[y]].value
#converts x and y points to a numpy array
x = np.array(x_points)
y = np.array(y_points)      
m, b = np.polyfit(x, y, 1)  #gets the slope and y-int of the linear regression line

plt.plot(x, y, 'o')         #plots the points in a scatter plot using dots

plt.plot(x, m*x + b, "r--", label=("y = " + str(m) + "x + " + str(b))) #plot lin-reg
plt.xlabel(str(input("Enter x-axis label: ")))
plt.ylabel(str(input("Enter y-axis label: ")))
plt.title(str(input("Enter graph title: ")))
print("\n\nLinear Regression Equation: y =", m, "* x +", b)
print("R value: ", getR(x_points, y_points))
#plt.axis([0,3000000,0,3000000])    #axis numberings
plt.legend()    #outputs a legend if lines have labels
plt.grid(True)  #turns on or off the grid
plt.show()      #outputs the plot
