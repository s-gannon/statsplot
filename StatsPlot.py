from matplotlib import pyplot as plt
import openpyxl as xl
import numpy as np

def resid(x_pts, y_pts, m, b):
    resid_pts = []
    for x in range(len(y_pts)):
        resid_pts.append(y_pts[x] - (m*x_pts[x] + b))
    return resid_pts

def getCorr(x_pts, y_pts):
    mean_x = mean_y = 0
    for x in range(len(x_pts)):
        mean_x += x_pts[x]
        mean_y += y_pts[x]
    mean_x = float(mean_x)/float(len(x_pts))
    mean_y = float(mean_y)/float(len(y_pts))
    a = x_pts
    b = y_pts
    for x in range(len(x_pts)):
        a[x] = mean_x - x_pts[x]
        b[x] = mean_y - y_pts[x]
    
x_points = []
y_points = []
x_range = ["",[0,0]]
y_range = ["",[0,0]]

print("Remember to put this file in the same folder as the Excel file!\n")
wb = xl.load_workbook(str(input("Enter in the Excel file name: ")))
ws = wb.active

x_start = str(input("Enter the start of the set of data for x-axis (format: A1): ")).upper()
x_end = str(input("Enter the end of the set of data for x-axis (format: A1): ")).upper()
y_start = str(input("Enter the start of the set of data for y-axis (format: A1): ")).upper()
y_end = str(input("Enter the end of the set of data for y-axis (format: A1): ")).upper()

x_range[0] = x_start[0]
x_start = x_start.replace(x_start[0], "")
x_end = x_end.replace(x_end[0], "")
x_range[1] = [int(x_start), int(x_end)]

y_range[0] = y_start[0]
y_start = y_start.replace(y_start[0], "")
y_end = y_end.replace(y_end[0], "")
y_range[1] = [int(y_start), int(y_end)]

for x in range(x_range[1][0], x_range[1][1] + 1):   #min start at 1, goes to n
    x_points.append(str(x_range[0] + str(x)))
    
for y in range(y_range[1][0], y_range[1][1] + 1):   #min start at 1, goes to n
    y_points.append(str(y_range[0] + str(y)))

for x in range(len(x_points)):
    x_points[x] = ws[x_points[x]].value

for y in range(len(y_points)):
    y_points[y] = ws[y_points[y]].value

x = np.array(x_points)
y = np.array(y_points)      #converts x and y points to a numpy array
plt.plot(x, y, 'o')         #plots the points in a scatter plot

m, b = np.polyfit(x, y, 1)  #gets the slope and y-int of the linear regression line

plt.plot(x, m*x + b, label=("y = " + str(m) + "x + " + str(b))) #plot lin-reg
plt.xlabel(str(input("Enter x-axis label: ")))
plt.ylabel(str(input("Enter y-axis label: ")))
plt.title(str(input("Enter graph title: ")))
print("\n\nLinear Regression Equation: y =", m, "* x +", b)

plt.legend()
plt.show()
